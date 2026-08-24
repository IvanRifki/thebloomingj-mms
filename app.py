from datetime import date, datetime
from zoneinfo import ZoneInfo

import base64
import csv
import io
import os
import re
import uuid
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.image import MIMEImage

from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build

import qrcode
from dotenv import load_dotenv

from flask import (
    Flask,
    Response,
    redirect,
    render_template,
    request,
    url_for,
    jsonify,
    send_file,
    flash,
    send_from_directory,
    make_response
)

from flask_login import (
    LoginManager,
    UserMixin,
    login_user,
    logout_user,
    login_required,
    current_user
)

from flask_sqlalchemy import SQLAlchemy

from supabase import create_client, Client

from werkzeug.security import generate_password_hash, check_password_hash

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle


load_dotenv()

MAIL_USERNAME = os.getenv("MAIL_USERNAME")
MAIL_SENDER_NAME = os.getenv("MAIL_SENDER_NAME", "Seminar Offline")

GMAIL_CLIENT_ID = os.getenv("GMAIL_CLIENT_ID")
GMAIL_CLIENT_SECRET = os.getenv("GMAIL_CLIENT_SECRET")
GMAIL_REFRESH_TOKEN = os.getenv("GMAIL_REFRESH_TOKEN")

GMAIL_SCOPES = [
    "https://www.googleapis.com/auth/gmail.send"
]

basedir = os.path.abspath(os.path.dirname(__file__))

app = Flask(__name__)

db_url = os.environ.get(
    'DATABASE_URL',
    'sqlite:///' + os.path.join(basedir, 'tiket.db')
)

if db_url.startswith('postgres://'):
    db_url = db_url.replace('postgres://', 'postgresql://', 1)

app.config['SQLALCHEMY_DATABASE_URI'] = db_url
app.config['SECRET_KEY'] = os.environ.get(
    'SECRET_KEY',
    'the-blooming-journey-secret-key-2026'
)
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

try:
    with app.app_context():
        db.create_all()
except Exception as e:
    print(f"[WARNING] db.create_all() gagal: {e}")

supabase_url = os.environ.get('SUPABASE_URL')
supabase_key = os.environ.get('SUPABASE_KEY')

supabase = create_client(
    supabase_url,
    supabase_key
)

WIB = ZoneInfo('Asia/Jakarta')


def kirim_email(penerima, subjek, isi_html, qr_bytes=None):
    creds = Credentials(
        token=None,
        refresh_token=GMAIL_REFRESH_TOKEN,
        token_uri="https://oauth2.googleapis.com/token",
        client_id=GMAIL_CLIENT_ID,
        client_secret=GMAIL_CLIENT_SECRET,
        scopes=GMAIL_SCOPES
    )

    service = build(
        "gmail",
        "v1",
        credentials=creds
    )

    msg = MIMEMultipart("related")

    msg["From"] = f"{MAIL_SENDER_NAME} <{MAIL_USERNAME}>"
    msg["To"] = penerima
    msg["Subject"] = subjek

    msg.attach(
        MIMEText(
            isi_html,
            "html",
            "utf-8"
        )
    )

    if qr_bytes:
        qr_image = MIMEImage(
            qr_bytes,
            _subtype="png"
        )

        qr_image.add_header(
            "Content-ID",
            "<qr_tiket>"
        )

        qr_image.add_header(
            "Content-Disposition",
            "inline",
            filename="qr-tiket.png"
        )

        msg.attach(qr_image)

    raw_message = base64.urlsafe_b64encode(
        msg.as_bytes()
    ).decode("utf-8")

    result = service.users().messages().send(
        userId="me",
        body={
            "raw": raw_message
        }
    ).execute()

    print("GMAIL MESSAGE ID:", result.get("id"))


def wib_now():
    return datetime.now(WIB)


def format_wib(dt):
    if dt is None:
        return '-'

    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=WIB)
    else:
        dt = dt.astimezone(WIB)

    return dt.strftime('%d/%m/%Y %H:%M WIB')


@app.template_filter('wib')
def wib_filter(dt):
    return format_wib(dt)


def upload_bukti_transfer(file):
    """Upload bukti transfer ke Supabase Storage."""

    if not file or not file.filename:
        raise ValueError(
            'Bukti transfer wajib diupload.'
        )

    max_size = 3 * 1024 * 1024

    file_bytes = file.read()

    if not file_bytes:
        raise ValueError(
            'File bukti transfer kosong.'
        )

    if len(file_bytes) > max_size:
        raise ValueError(
            'Ukuran bukti transfer maksimal 3 MB.'
        )

    filename_asli = file.filename

    if '.' not in filename_asli:
        raise ValueError(
            'File bukti transfer tidak memiliki ekstensi.'
        )

    ext = filename_asli.rsplit(
        '.',
        1
    )[-1].lower()

    allowed_extensions = {
        'jpg',
        'jpeg',
        'png',
        'pdf'
    }

    if ext not in allowed_extensions:
        raise ValueError(
            'Format bukti transfer harus JPG, PNG, JPEG, atau PDF.'
        )

    filename = f"{uuid.uuid4().hex}.{ext}"

    content_type_map = {
        'jpg': 'image/jpeg',
        'jpeg': 'image/jpeg',
        'png': 'image/png',
        'pdf': 'application/pdf'
    }

    content_type = content_type_map[ext]

    try:
        bucket = supabase.storage.from_(
            'bukti-transfer'
        )

        bucket.upload(
            filename,
            file_bytes,
            file_options={
                'content-type': content_type,
                'cache-control': '3600',
                'upsert': False
            }
        )

        public_url = bucket.get_public_url(
            filename
        )

        print(
            'UPLOAD BUKTI BERHASIL:',
            public_url
        )

        return public_url

    except Exception as e:
        print(
            'SUPABASE STORAGE ERROR:',
            repr(e)
        )
        raise


def sanitize_input(text, max_length=100):
    if not text:
        return ''

    cleaned = re.sub(r'<[^>]+>', '', text)
    cleaned = re.sub(
        r'[\x00-\x08\x0b\x0c\x0e-\x1f]',
        '',
        cleaned
    )

    return cleaned.strip()[:max_length]


login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Silakan login terlebih dahulu.'
login_manager.login_message_category = 'warning'


class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(200), nullable=False)

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(
            self.password_hash,
            password
        )


class Tiket(db.Model):
    __tablename__ = 'tiket'

    id = db.Column(db.Integer, primary_key=True)
    kode = db.Column(
        db.String(10),
        unique=True,
        nullable=False
    )
    nama = db.Column(db.String(100))
    no_telp = db.Column(db.String(20))
    email = db.Column(db.String(100))
    kode_referal = db.Column(db.String(50))
    jenis_tiket = db.Column(db.String(50))
    jumlah_peserta = db.Column(
        db.Integer,
        default=1
    )
    is_used = db.Column(
        db.Boolean,
        default=False
    )
    waktu_daftar = db.Column(
        db.DateTime,
        default=wib_now
    )
    waktu_scan = db.Column(
        db.DateTime,
        nullable=True
    )
    bukti_transfer = db.Column(
        db.String(255),
        nullable=True
    )
    bukti_terverifikasi = db.Column(
        db.Boolean,
        default=False
    )


class Setting(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    key = db.Column(
        db.String(50),
        unique=True,
        nullable=False
    )
    value = db.Column(
        db.String(200),
        nullable=False
    )


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


def get_kuota():
    setting = Setting.query.filter_by(
        key='kuota'
    ).first()

    return int(setting.value) if setting else 120


def set_kuota(kuota_baru):
    setting = Setting.query.filter_by(
        key='kuota'
    ).first()

    if setting:
        setting.value = str(kuota_baru)
    else:
        setting = Setting(
            key='kuota',
            value=str(kuota_baru)
        )
        db.session.add(setting)

    db.session.commit()


def get_limit_early_bird():
    setting = Setting.query.filter_by(
        key='limit_early_bird'
    ).first()

    return int(setting.value) if setting else 40


def set_limit_early_bird(limit_baru):
    setting = Setting.query.filter_by(
        key='limit_early_bird'
    ).first()

    if setting:
        setting.value = str(limit_baru)
    else:
        setting = Setting(
            key='limit_early_bird',
            value=str(limit_baru)
        )
        db.session.add(setting)

    db.session.commit()


def get_limit_normal():
    setting = Setting.query.filter_by(
        key='limit_normal'
    ).first()

    return int(setting.value) if setting else 220


def generate_qr_base64(data):
    img = qrcode.make(data)

    buffer = io.BytesIO()
    img.save(buffer, format='PNG')
    buffer.seek(0)

    return base64.b64encode(
        buffer.getvalue()
    ).decode('utf-8')


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/api/sisa_eb')
def api_sisa_eb():
    limit_early_bird = get_limit_early_bird()

    tiket_eb = Tiket.query.filter(
        Tiket.jenis_tiket.in_([
            'single_eb',
            'circle_eb',
            'squad_eb'
        ])
    ).all()

    total_early_bird = sum(
        t.jumlah_peserta for t in tiket_eb
    )

    sisa_eb = max(0, limit_early_bird - total_early_bird)

    resp = jsonify({
        'sisa_eb': sisa_eb,
        'eb_penuh': total_early_bird >= limit_early_bird
    })
    resp.headers['Cache-Control'] = 'no-store'
    return resp


@app.route('/daftar', methods=['GET', 'POST'])
def pendaftaran():

    kuota = get_kuota()

    total_terdaftar = db.session.query(
        db.func.sum(Tiket.jumlah_peserta)
    ).scalar() or 0

    if total_terdaftar >= kuota:
        return redirect('/habis')

    limit_early_bird = get_limit_early_bird()

    tiket_eb = Tiket.query.filter(
        Tiket.jenis_tiket.in_([
            'single_eb',
            'circle_eb',
            'squad_eb'
        ])
    ).all()

    total_early_bird = sum(
        t.jumlah_peserta for t in tiket_eb
    )

    if total_early_bird >= limit_early_bird:
        return redirect('/normal_daftar')

    if request.method == 'POST':

        total_terdaftar = db.session.query(
            db.func.sum(Tiket.jumlah_peserta)
        ).scalar() or 0

        if total_terdaftar >= kuota:
            return redirect('/habis')

        nama = sanitize_input(
            request.form.get('nama')
        )

        no_telp = sanitize_input(
            request.form.get('telp')
        )

        email = sanitize_input(
            request.form.get('email')
        )

        kode_referal = sanitize_input(
            request.form.get('ref')
        )

        if not nama:
            flash(
                'Nama wajib diisi!',
                'danger'
            )
            return redirect('/daftar')

        jenis_tiket = request.form.get(
            'paket_tiket'
        )

        map_peserta = {
            'single_eb': 1,
            'circle_eb': 3,
            'squad_eb': 5
        }

        peserta_baru = map_peserta.get(
            jenis_tiket,
            1
        )

        if total_terdaftar + peserta_baru > kuota:
            return redirect('/habis')

        if jenis_tiket in [
            'single_eb',
            'circle_eb',
            'squad_eb'
        ]:

            tiket_eb = Tiket.query.filter(
                Tiket.jenis_tiket.in_([
                    'single_eb',
                    'circle_eb',
                    'squad_eb'
                ])
            ).all()

            total_early_bird = sum(
                t.jumlah_peserta for t in tiket_eb
            )

            if (
                total_early_bird + peserta_baru
                > limit_early_bird
            ):
                flash(
                    'Mohon maaf, kuota khusus Early Bird sudah penuh! '
                    'Harga tiket kembali normal. Silahkan mengisi form kembali. '
                    'Gunakan kode referal untuk mendapatkan potongan harga (opsional).',
                    'eb-penuh'
                )

                return redirect(
                    '/normal_daftar?status=eb_penuh'
                )

        try:
            file_bukti = request.files.get('bukti')

            print(
                'FILE BUKTI:',
                file_bukti.filename
                if file_bukti and file_bukti.filename
                else 'TIDAK ADA FILE'
            )

            if file_bukti:
                print(
                    'CONTENT TYPE:',
                    file_bukti.content_type
                )

            bukti_url = upload_bukti_transfer(
                file_bukti
            )

        except ValueError as e:
            print(
                'UPLOAD VALIDATION ERROR:',
                repr(e)
            )

            flash(
                str(e),
                'danger'
            )

            return redirect('/daftar')

        except Exception as e:
            print(
                'UPLOAD BUKTI ERROR:',
                repr(e)
            )

            db.session.rollback()

            flash(
                f'Upload bukti transfer gagal: {str(e)}',
                'danger'
            )

            return redirect('/daftar')

        kode = "MMS-" + str(
            uuid.uuid4()
        ).upper()[:4]

        tiket_baru = Tiket(
            kode=kode,
            nama=nama,
            no_telp=no_telp,
            email=email,
            kode_referal=kode_referal,
            jenis_tiket=jenis_tiket,
            jumlah_peserta=peserta_baru,
            waktu_daftar=wib_now(),
            bukti_transfer=bukti_url
        )

        db.session.add(tiket_baru)
        db.session.commit()

        qr_base64 = generate_qr_base64(kode)

        return render_template(
            'sukses.html',
            nama=nama,
            kode=kode,
            angkatan=jenis_tiket,
            qr_base64=qr_base64,
            waktu_daftar=format_wib(
                wib_now()
            )
        )

    sisa_eb = max(
        0,
        limit_early_bird - total_early_bird
    )

    resp = make_response(render_template(
        'daftar.html',
        tiket_terjual=total_early_bird,
        sisa_eb=sisa_eb
    ))
    resp.headers['Cache-Control'] = 'no-store'
    return resp


@app.route('/normal_daftar', methods=['GET', 'POST'])
def pendaftaran_normal():

    KODE_REFERAL_VALID = [
        'MMS2026',
        'SAHABATMMS',
        'BLOOMING01'
    ]

    kuota = get_kuota()

    total_terdaftar = db.session.query(
        db.func.sum(Tiket.jumlah_peserta)
    ).scalar() or 0

    if total_terdaftar >= kuota:
        return redirect('/habis')

    limit_normal = get_limit_normal()

    tiket_normal = Tiket.query.filter_by(
        jenis_tiket='normal'
    ).all()

    total_normal = sum(
        t.jumlah_peserta for t in tiket_normal
    )

    if total_normal >= limit_normal:
        return redirect('/habis')

    if request.method == 'POST':

        total_terdaftar = db.session.query(
            db.func.sum(Tiket.jumlah_peserta)
        ).scalar() or 0

        if total_terdaftar >= kuota:
            return redirect('/habis')

        nama = sanitize_input(
            request.form.get('nama')
        )

        no_telp = sanitize_input(
            request.form.get('telp')
        )

        email = sanitize_input(
            request.form.get('email')
        )

        kode_referal = sanitize_input(
            request.form.get('ref')
        )

        if not nama:
            flash(
                'Nama wajib diisi!',
                'danger'
            )
            return redirect('/normal_daftar')

        jenis_tiket = 'normal'
        peserta_baru = 1

        if total_terdaftar + peserta_baru > kuota:
            return redirect('/habis')

        tiket_normal = Tiket.query.filter_by(
            jenis_tiket='normal'
        ).all()

        total_normal = sum(
            t.jumlah_peserta for t in tiket_normal
        )

        if total_normal + peserta_baru > limit_normal:
            return redirect('/habis')

        kode_ref_upper = (
            kode_referal or ''
        ).strip().upper()

        if kode_ref_upper in KODE_REFERAL_VALID:
            harga_final = 398000
            kode_referal_final = kode_ref_upper
        else:
            harga_final = 399000
            kode_referal_final = (
                kode_referal or None
            )

        try:
            file_bukti = request.files.get('bukti')

            print(
                'FILE BUKTI:',
                file_bukti.filename
                if file_bukti and file_bukti.filename
                else 'TIDAK ADA FILE'
            )

            if file_bukti:
                print(
                    'CONTENT TYPE:',
                    file_bukti.content_type
                )

            bukti_url = upload_bukti_transfer(
                file_bukti
            )

        except ValueError as e:
            print(
                'UPLOAD VALIDATION ERROR:',
                repr(e)
            )

            flash(
                str(e),
                'danger'
            )

            return redirect('/normal_daftar')

        except Exception as e:
            print(
                'UPLOAD BUKTI ERROR:',
                repr(e)
            )

            db.session.rollback()

            flash(
                f'Upload bukti transfer gagal: {str(e)}',
                'danger'
            )

            return redirect('/normal_daftar')

        kode = "MMS-" + str(
            uuid.uuid4()
        ).upper()[:4]

        tiket_baru = Tiket(
            kode=kode,
            nama=nama,
            no_telp=no_telp,
            email=email,
            kode_referal=kode_referal_final,
            jenis_tiket='normal',
            jumlah_peserta=1,
            waktu_daftar=wib_now(),
            bukti_transfer=bukti_url
        )

        db.session.add(tiket_baru)
        db.session.commit()

        return redirect(
            url_for(
                'sukses',
                kode=kode
            )
        )

    sisa_normal = max(
        0,
        limit_normal - total_normal
    )

    return render_template(
        'normal_daftar.html',
        sisa_normal=sisa_normal
    )


@app.route('/test-email')
def test_email():
    try:
        kirim_email(
            "madrasahwanita21@gmail.com",
            "Tes Email - Seminar Offline",
            """
            <html>
                <body>
                    <h2>Halo!</h2>
                    <p>Ini adalah email percobaan dari website <b>Seminar Offline</b>.</p>
                    <p>Kalau kamu menerima email ini, berarti Gmail SMTP sudah berhasil 🎉</p>
                </body>
            </html>
            """
        )

        return "Email berhasil dikirim!"

    except Exception as e:
        return f"Email gagal dikirim: {e}", 500


@app.route('/sukses')
def sukses():

    kode = request.args.get('kode')

    if not kode:
        flash(
            'Data pendaftaran tidak ditemukan.',
            'danger'
        )
        return redirect('/normal_daftar')

    tiket = Tiket.query.filter_by(
        kode=kode
    ).first()

    if not tiket:
        flash(
            'Data tiket tidak ditemukan.',
            'danger'
        )
        return redirect('/normal_daftar')

    qr_base64 = generate_qr_base64(kode)

    return render_template(
        'sukses.html',
        nama=tiket.nama,
        kode=tiket.kode,
        angkatan=tiket.jenis_tiket,
        qr_base64=qr_base64,
        waktu_daftar=format_wib(
            tiket.waktu_daftar
        )
    )


@app.route('/hapus_tiket/<int:tiket_id>', methods=['POST'])
def hapus_tiket(tiket_id):

    tiket = Tiket.query.get(tiket_id)

    is_ajax = (
        request.headers.get(
            'X-Requested-With'
        ) == 'XMLHttpRequest'
    )

    if not tiket:

        if is_ajax:
            return jsonify({
                'success': False,
                'message': 'Data tidak ditemukan'
            }), 404

        flash(
            'Data tidak ditemukan',
            'danger'
        )

        return redirect('/admin')

    db.session.delete(tiket)
    db.session.commit()

    if is_ajax:
        return jsonify({
            'success': True
        })

    flash(
        'Data berhasil dihapus',
        'success'
    )

    return redirect('/admin')


@app.route('/hapus_semua_peserta', methods=['POST'])
def hapus_semua_peserta():

    Tiket.query.delete()
    db.session.commit()

    flash(
        'Semua data peserta berhasil dihapus',
        'success'
    )

    return redirect('/admin')


@app.route('/reset_semua', methods=['POST'])
def reset_semua():

    Tiket.query.update({
        Tiket.is_used: False
    })

    db.session.commit()

    flash(
        'Semua status kehadiran berhasil direset menjadi Belum Hadir.',
        'success'
    )

    return redirect('/admin')


@app.route('/export/excel')
def export_excel():

    tiket = Tiket.query.order_by(
        Tiket.id.asc()
    ).all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Data Peserta'

    headers = [
        'No', 'Kode Tiket', 'Nama Lengkap', 'No Telp', 'Email',
        'Kode Referal', 'Jenis Tiket', 'Jumlah Peserta',
        'Status', 'Waktu Daftar', 'Waktu Scan'
    ]

    ws.append(headers)

    header_fill = PatternFill(
        start_color='732D3A',
        end_color='732D3A',
        fill_type='solid'
    )

    for cell in ws[1]:
        cell.font = Font(color='FFFFFF', bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for no, t in enumerate(tiket, start=1):
        ws.append([
            no,
            t.kode or '-',
            t.nama or '-',
            t.no_telp or '-',
            t.email or '-',
            t.kode_referal or '-',
            t.jenis_tiket or '-',
            t.jumlah_peserta or 0,
            'Hadir' if t.is_used else 'Belum Hadir',
            format_wib(t.waktu_daftar) if t.waktu_daftar else '-',
            format_wib(t.waktu_scan) if t.waktu_scan else '-'
        ])

    column_widths = [6, 14, 22, 16, 26, 16, 14, 14, 14, 20, 20]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[
            ws.cell(row=1, column=i).column_letter
        ].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition':
                'attachment; filename=data_peserta_mms.xlsx'
        }
    )


@app.route('/export/pdf')
def export_pdf():

    tiket = Tiket.query.order_by(
        Tiket.id.asc()
    ).all()

    output = io.BytesIO()

    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        topMargin=20,
        bottomMargin=20
    )

    data = [[
        'No', 'Kode', 'Nama', 'No Telp', 'Email',
        'Referal', 'Jenis', 'Peserta', 'Status', 'Daftar', 'Scan'
    ]]

    for no, t in enumerate(tiket, start=1):
        data.append([
            str(no),
            t.kode or '-',
            t.nama or '-',
            t.no_telp or '-',
            t.email or '-',
            t.kode_referal or '-',
            t.jenis_tiket or '-',
            str(t.jumlah_peserta or 0),
            'Hadir' if t.is_used else 'Belum Hadir',
            format_wib(t.waktu_daftar) if t.waktu_daftar else '-',
            format_wib(t.waktu_scan) if t.waktu_scan else '-'
        ])

    table = Table(data, repeatRows=1)

    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#732D3A')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E5D5D0')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1),
            [colors.white, colors.HexColor('#FAF6F5')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))

    doc.build([table])

    output.seek(0)

    return Response(
        output.getvalue(),
        mimetype='application/pdf',
        headers={
            'Content-Disposition':
                'attachment; filename=data_peserta_mms.pdf'
        }
    )


@app.route('/hadirkan_manual', methods=['POST'])
def hadirkan_manual():

    data = request.get_json()
    tiket_id = data.get('id')

    if not tiket_id:
        return jsonify({
            'success': False,
            'message': 'ID tiket tidak ditemukan.'
        }), 400

    tiket = Tiket.query.get(tiket_id)

    if not tiket:
        return jsonify({
            'success': False,
            'message': 'Tiket tidak ditemukan.'
        }), 404

    tiket.is_used = True
    tiket.waktu_scan = wib_now()

    db.session.commit()

    return jsonify({
        'success': True,
        'message': 'Peserta berhasil dihadirkan.'
    })


@app.route('/verify_bukti/<int:tiket_id>', methods=['POST'])
@login_required
def verify_bukti(tiket_id):
    try:
        tiket = Tiket.query.get_or_404(tiket_id)

        data = request.get_json() or {}
        verified_baru = data.get('verified', False)

        sebelumnya_terverifikasi = tiket.bukti_terverifikasi

        print("=== VERIFY BUKTI ===")
        print("Tiket ID:", tiket.id)
        print("Kode:", tiket.kode)
        print("Email penerima:", tiket.email)
        print("Sebelumnya:", sebelumnya_terverifikasi)
        print("Sekarang:", verified_baru)

        tiket.bukti_terverifikasi = verified_baru

        if verified_baru and not sebelumnya_terverifikasi:
            print("STATUS BERUBAH MENJADI TERVERIFIKASI")
            print("Mencoba mengirim email ke:", tiket.email)

            qr = qrcode.make(tiket.kode)

            qr_buffer = io.BytesIO()
            qr.save(qr_buffer, format='PNG')
            qr_buffer.seek(0)

            qr_bytes = qr_buffer.getvalue()

            whatsapp_group_link = os.getenv(
                "WHATSAPP_GROUP_LINK",
                ""
            )

            isi_email = f"""
            <html>
                <body style="
                    margin:0;
                    padding:30px;
                    background:#f8f5f3;
                    font-family:Arial, sans-serif;
                    color:#333;
                ">
                    <div style="
                        max-width:600px;
                        margin:auto;
                        background:#ffffff;
                        padding:30px;
                        border-radius:16px;
                    ">

                        <h2 style="text-align:center;">
                            Alhamdulillah, pembayaran Berhasil Diverifikasi 🎉
                        </h2>

                        <p>
                            Halo <b>{tiket.nama}</b>,
                        </p>

                        <p>
                            Pembayaran kamu untuk
                            <b>The Blooming Journey 2026</b>
                            telah berhasil diverifikasi oleh panitia.
                        </p>

                        <hr>

                        <p>
                            <b>Kode Tiket:</b><br>
                            {tiket.kode}
                        </p>

                        <p>
                            <b>Jenis Tiket:</b><br>
                            {tiket.jenis_tiket or '-'}
                        </p>

                        <div style="
                            text-align:center;
                            margin:30px 0;
                        ">
                            <p>
                                <b>QR Code Tiket Kamu</b>
                            </p>

                            <img
                                src="cid:qr_tiket"
                                alt="QR Code Tiket"
                                style="
                                    width:220px;
                                    height:220px;
                                "
                            >

                            <p style="
                                font-size:13px;
                                color:#777;
                            ">
                                Simpan QR Code ini dan
                                tunjukkan saat registrasi acara.
                            </p>
                        </div>

                        <hr>

                        <div style="text-align:center;">
                            <p>
                                <b>Gabung ke Grup WhatsApp Peserta</b>
                            </p>

                            <a
                                href="{whatsapp_group_link}"
                                style="
                                    display:inline-block;
                                    padding:12px 22px;
                                    background:#25D366;
                                    color:white;
                                    text-decoration:none;
                                    border-radius:8px;
                                    font-weight:bold;
                                "
                            >
                                Gabung Grup WhatsApp
                            </a>
                        </div>

                        <p style="
                            margin-top:30px;
                            text-align:center;
                        ">
                            Sampai bertemu di acara! 🤍
                        </p>

                        <p>
                            With Love,<br>
                            <b>The Blooming Journey Crew</b>
                        </p>

                    </div>
                </body>
            </html>
            """

            print("Memanggil kirim_email()...")

            kirim_email(
                tiket.email,
                "Pembayaran Terverifikasi - The Blooming Journey 2026",
                isi_email,
                qr_bytes
            )

            print("EMAIL BERHASIL DIKIRIM!")

        db.session.commit()

        print("DATABASE BERHASIL DI-COMMIT")
        print("========================")

        return jsonify({
            "success": True
        })

    except Exception as e:
        db.session.rollback()

        print("=== ERROR VERIFY BUKTI ===")
        print(repr(e))
        print("==========================")

        return jsonify({
            "success": False,
            "message": str(e)
        }), 500


@app.route('/admin/kuota', methods=['POST'])
def admin_kuota():

    try:
        jenis = request.form.get('jenis')

        if jenis == 'early_bird':
            limit_baru = int(
                request.form.get(
                    'limit_early_bird',
                    0
                )
            )

            if limit_baru < 1:
                flash(
                    'Kuota Early Bird minimal 1.',
                    'danger'
                )
                return redirect(
                    url_for('admin')
                )

            set_limit_early_bird(limit_baru)

            flash(
                f'Kuota Early Bird berhasil diubah menjadi {limit_baru}.',
                'success'
            )

        else:
            kuota_baru = int(
                request.form.get(
                    'kuota',
                    0
                )
            )

            if kuota_baru < 1:
                flash(
                    'Kuota minimal 1.',
                    'danger'
                )
                return redirect(
                    url_for('admin')
                )

            set_kuota(kuota_baru)

            flash(
                f'Kuota berhasil diubah menjadi {kuota_baru}.',
                'success'
            )

    except (ValueError, TypeError):

        flash(
            'Nilai kuota tidak valid.',
            'danger'
        )

    return redirect(
        url_for('admin')
    )


@app.route('/login', methods=['GET', 'POST'])
def login():

    if current_user.is_authenticated:
        return redirect(
            url_for('admin')
        )

    if request.method == 'POST':

        username = request.form.get(
            'username'
        )

        password = request.form.get(
            'password'
        )

        user = User.query.filter_by(
            username=username
        ).first()

        if user and user.check_password(
            password
        ):
            login_user(user)

            next_page = request.args.get(
                'next'
            )

            return redirect(
                next_page or url_for('admin')
            )

        flash(
            'Username atau password salah!',
            'danger'
        )

    return render_template(
        'login.html'
    )


@app.route('/logout')
@login_required
def logout():

    logout_user()

    flash(
        'Berhasil logout.',
        'info'
    )

    return redirect(
        url_for('login')
    )


@app.route('/admin')
@login_required
def admin():

    try:
        tiket = Tiket.query.all()

        total = len(tiket)

        terpakai = sum(
            1 for t in tiket
            if t.is_used
        )

        sisa = total - terpakai

        kuota = get_kuota()
        limit_early_bird = get_limit_early_bird()

        return render_template(
            'admin.html',
            tiket=tiket,
            total=total,
            terpakai=terpakai,
            sisa=sisa,
            kuota=kuota,
            limit_early_bird=limit_early_bird
        )

    except Exception:
        return (
            'Terjadi kesalahan saat membuka dashboard admin.',
            500
        )


@app.route('/habis')
def habis():

    kuota = get_kuota()

    total_terdaftar = db.session.query(
        db.func.sum(Tiket.jumlah_peserta)
    ).scalar() or 0

    return render_template(
        'habis.html',
        kuota=kuota,
        total=total_terdaftar
    )


@app.route('/scan')
def scan():

    return render_template(
        'scan.html',
        status=None
    )


@app.route('/scan/<kode_tiket>')
def scan_kode(kode_tiket):

    tiket = Tiket.query.filter_by(
        kode=kode_tiket
    ).first()

    if not tiket:

        return render_template(
            'scan.html',
            status='tidak ditemukan',
            kode=kode_tiket
        )

    if tiket.is_used:

        return render_template(
            'scan.html',
            status='telah terpakai',
            kode=kode_tiket,
            nama_peserta=tiket.nama,
            angkatan_peserta=tiket.jenis_tiket
        )

    tiket.is_used = True
    tiket.waktu_scan = wib_now()

    db.session.commit()

    return render_template(
        'scan.html',
        status='berhasil',
        kode=kode_tiket,
        nama_peserta=tiket.nama,
        angkatan_peserta=tiket.jenis_tiket
    )


if __name__ == '__main__':
    app.run(debug=True)
